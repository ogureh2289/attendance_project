import threading
import tkinter as tk
from datetime import datetime
from tkinter import ttk, messagebox

from controller import create_student
from db import get_attendance
from scanner import start_qr_scanner, run_scanner_in_thread, stop_and_join_scanner, mark_absent_students
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

class MainApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Система посещаемости")
        self.root.geometry("400x300")
        self.error_shown = False
        self.create_main_menu()

    def create_main_menu(self):
        self.clear_window()

        title_label = tk.Label(self.root, text="Главное меню", font=("Arial", 16))
        title_label.pack(pady=20)

        btn_lesson = tk.Button(self.root, text="Урок", command=self.open_lesson_window, width=20, height=2)
        btn_lesson.pack(pady=10)

        btn_add = tk.Button(self.root, text="Добавить", command=self.open_add_window, width=20, height=2)
        btn_add.pack(pady=10)

        btn_print = tk.Button(self.root, text="Печать", command=self.open_print_window, width=20, height=2)
        btn_print.pack(pady=10)

    def open_lesson_window(self):
        self.clear_window()

        title_label = tk.Label(self.root, text="Посещаемость", font=("Arial", 16))
        title_label.pack(pady=10)

        frame = tk.Frame(self.root)
        frame.pack(pady=20)

        tk.Label(frame, text="Урок:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.lesson_name = tk.Entry(frame)
        self.lesson_name.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(frame, text="Дата (ДД.ММ.ГГГГ):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.lesson_date = tk.Entry(frame)
        self.lesson_date.grid(row=1, column=1, padx=5, pady=5)
        self.lesson_date.insert(0, datetime.now().strftime("%d.%m.%Y"))

        tk.Label(frame, text="Время начала:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.lesson_start = tk.Entry(frame)
        self.lesson_start.grid(row=2, column=1, padx=5, pady=5)
        self.lesson_start.insert(0, datetime.now().strftime("%H:%M"))

        tk.Label(frame, text="Класс:").grid(row=3, column=0, padx=5, pady=5, sticky="e")
        self.student_class = tk.Entry(frame)
        self.student_class.grid(row=3, column=1, padx=5, pady=5)

        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=20)

        btn_start = tk.Button(btn_frame, text="Старт", command=self.start_lesson, width=15)
        btn_start.pack(side=tk.LEFT, padx=10)

        btn_back = tk.Button(btn_frame, text="Назад", command=self.end_lesson, width=15)
        btn_back.pack(side=tk.RIGHT, padx=10)

    def start_lesson(self):
        if not all([self.lesson_name.get(), self.lesson_date.get(),
                    self.lesson_start.get(), self.student_class.get()]):
            messagebox.showerror("Ошибка", "Все поля должны быть заполнены!")
            return

        lesson_name = self.lesson_name.get()
        lesson_date = self.lesson_date.get()
        lesson_start = self.lesson_start.get()
        student_class = self.student_class.get()

        self.clear_window()
        tk.Label(self.root, text="Окно обработки посещаемости", font=("Arial", 14)).pack(pady=20)

        stop_button = tk.Button(self.root, text="Завершить урок", command=self.end_lesson, width=20)
        stop_button.pack(pady=10)

        # Запускаем сканер в отдельном потоке
        threading.Thread(
            target=start_qr_scanner,
            args=(lesson_name, lesson_date, lesson_start, student_class),
            kwargs={
                'on_success': self.show_success_messagebox,
                'on_failure': self.safe_show_error_messagebox
            },
            daemon=True
        ).start()

    def end_lesson(self):
        global scanner_active
        scanner_active = False
        stop_and_join_scanner()  # подождёт завершения потока
        mark_absent_students()  # записывает отсутствующих
        self.create_main_menu()  # возвращение в главное меню

    def start_camera(self, lesson_name, lesson_date, lesson_start, student_class):
        self.error_shown = False
        while self.scanner_active:
            run_scanner_in_thread(
                lesson_name=self.lesson_name,
                lesson_date=self.lesson_date,
                lesson_start=self.lesson_start,
                student_class=self.student_class,
                on_success=self.on_success_message,
                on_failure=self.on_failure_message
            )

    def show_success_messagebox(self, msg):
        if not self.error_shown:
            self.error_shown = True

            def show_and_reset():
                messagebox.showinfo("Успех", msg)
                self.error_shown = False

            self.root.after(0, show_and_reset)

    def safe_show_error_messagebox(self, msg):
        if not self.error_shown:
            self.error_shown = True

            def show_and_reset():
                messagebox.showerror("Ошибка", msg)
                self.error_shown = False

            self.root.after(0, show_and_reset)

    def open_add_window(self):
        self.clear_window()

        title_label = tk.Label(self.root, text="Добавить ученика", font=("Arial", 16))
        title_label.pack(pady=10)

        frame = tk.Frame(self.root)
        frame.pack(pady=20)

        tk.Label(frame, text="ФИО:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.full_name = tk.Entry(frame)
        self.full_name.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(frame, text="Класс:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.add_class = tk.Entry(frame)
        self.add_class.grid(row=1, column=1, padx=5, pady=5)

        tk.Label(frame, text="Email:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.email = tk.Entry(frame)
        self.email.grid(row=2, column=1, padx=5, pady=5)

        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=20)

        btn_ok = tk.Button(btn_frame, text="OK", command=self.add_student, width=15)
        btn_ok.pack(side=tk.LEFT, padx=10)

        btn_back = tk.Button(btn_frame, text="Назад", command=self.create_main_menu, width=15)
        btn_back.pack(side=tk.RIGHT, padx=10)

    def add_student(self):

        if not all([self.full_name.get(), self.add_class.get(), self.email.get()]):
            messagebox.showerror("Ошибка", "Все поля должны быть заполнены!")
            return

        try:
            create_student(self.full_name.get(), self.add_class.get(), self.email.get())
            messagebox.showinfo("Успех", "Ученик успешно добавлен!")
            self.create_main_menu()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")

    def open_print_window(self):
        self.clear_window()

        title_label = tk.Label(self.root, text="Печать данных", font=("Arial", 16))
        title_label.pack(pady=10)

        frame = tk.Frame(self.root)
        frame.pack(pady=20)

        tk.Label(frame, text="Класс:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.print_class = tk.Entry(frame)
        self.print_class.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(frame, text="Дата (ДД.ММ.ГГГГ):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.print_date = tk.Entry(frame)
        self.print_date.grid(row=1, column=1, padx=5, pady=5)
        self.print_date.insert(0, datetime.now().strftime("%d.%m.%Y"))

        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=20)

        btn_ok = tk.Button(btn_frame, text="OK", command=self.print_data, width=15)
        btn_ok.pack(side=tk.LEFT, padx=10)

        btn_back = tk.Button(btn_frame, text="Назад", command=self.create_main_menu, width=15)
        btn_back.pack(side=tk.RIGHT, padx=10)

    def show_data_window(self, data, class_name, date):
        self.clear_window()

        title_label = tk.Label(self.root, text=f"Данные посещаемости\nКласс: {class_name}, Дата: {date}",
                               font=("Arial", 14))
        title_label.pack(pady=10)

        frame = tk.Frame(self.root)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Обновленные колонки с дополнительной информацией
        tree = ttk.Treeview(frame, columns=("id", "name", "class", "lesson", "date", "start", "arrival", "status"),
                            show="headings")

        # Настройка заголовков
        tree.heading("id", text="ID")
        tree.heading("name", text="ФИО")
        tree.heading("class", text="Класс")
        tree.heading("lesson", text="Урок")
        tree.heading("date", text="Дата")
        tree.heading("start", text="Начало")
        tree.heading("arrival", text="Приход")
        tree.heading("status", text="Статус")

        # Настройка ширины колонок
        tree.column("id", width=40, anchor=tk.CENTER)
        tree.column("name", width=150)
        tree.column("class", width=60, anchor=tk.CENTER)
        tree.column("lesson", width=100)
        tree.column("date", width=80)
        tree.column("start", width=60, anchor=tk.CENTER)
        tree.column("arrival", width=60, anchor=tk.CENTER)
        tree.column("status", width=100, anchor=tk.CENTER)

        # Заполнение данными
        for record in data:
            tree.insert("", tk.END, values=record)

        # Добавление прокрутки
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(fill=tk.BOTH, expand=True)

        # Кнопка назад
        btn_back = tk.Button(self.root, text="Назад", command=self.create_main_menu, width=15)
        btn_back.pack(pady=10)

    def print_data(self):
        if not all([self.print_class.get(), self.print_date.get()]):
            messagebox.showerror("Ошибка", "Все поля должны быть заполнены!")
            return

        class_name = self.print_class.get()
        date = self.print_date.get()

        try:
            # Используем новую функцию для получения данных
            students = get_attendance(date, class_name)
            if not students:
                messagebox.showinfo("Информация", "Нет данных для отображения")
                return

            self.show_data_window(students, class_name, date)
        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")

    def clear_window(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def export_to_excel(self, data, class_name, date):
        # Создаем новую книгу Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Посещаемость"

        # Заголовки
        headers = ["ID", "ФИО", "Класс", "Урок", "Дата", "Начало", "Приход", "Статус"]
        ws.append(headers)

        # Стили для заголовков
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Добавляем данные
        for row in data:
            ws.append(row)

        # Настраиваем ширину столбцов
        column_widths = {
            'A': 8,  # ID
            'B': 30,  # ФИО
            'C': 10,  # Класс
            'D': 20,  # Урок
            'E': 12,  # Дата
            'F': 10,  # Начало
            'G': 10,  # Приход
            'H': 15  # Статус
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Центрируем данные в ячейках
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center')

        # Сохраняем файл
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Посещаемость_{class_name}_{date.replace('.', '-')}_{timestamp}.xlsx"

        # Создаем папку exports, если ее нет
        if not os.path.exists('exports'):
            os.makedirs('exports')

        filepath = os.path.join('exports', filename)
        wb.save(filepath)
        return filepath

    # Модифицируем метод show_data_window
    def show_data_window(self, data, class_name, date):
        self.clear_window()

        title_label = tk.Label(self.root, text=f"Данные посещаемости\nКласс: {class_name}, Дата: {date}",
                               font=("Arial", 14))
        title_label.pack(pady=10)

        frame = tk.Frame(self.root)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Обновленные колонки с дополнительной информацией
        tree = ttk.Treeview(frame, columns=("id", "name", "class", "lesson", "date", "start", "arrival", "status"),
                            show="headings")

        # Настройка заголовков
        tree.heading("id", text="ID")
        tree.heading("name", text="ФИО")
        tree.heading("class", text="Класс")
        tree.heading("lesson", text="Урок")
        tree.heading("date", text="Дата")
        tree.heading("start", text="Начало")
        tree.heading("arrival", text="Приход")
        tree.heading("status", text="Статус")

        # Настройка ширины колонок
        tree.column("id", width=40, anchor=tk.CENTER)
        tree.column("name", width=150)
        tree.column("class", width=60, anchor=tk.CENTER)
        tree.column("lesson", width=100)
        tree.column("date", width=80)
        tree.column("start", width=60, anchor=tk.CENTER)
        tree.column("arrival", width=60, anchor=tk.CENTER)
        tree.column("status", width=100, anchor=tk.CENTER)

        # Заполнение данными
        for record in data:
            tree.insert("", tk.END, values=record)

        # Добавление прокрутки
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(fill=tk.BOTH, expand=True)

        # Фрейм для кнопок
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=10)

        # Кнопка экспорта в Excel
        btn_export = tk.Button(btn_frame, text="Экспорт в Excel",
                               command=lambda: self.export_and_show_message(data, class_name, date),
                               width=15)
        btn_export.pack(side=tk.LEFT, padx=10)

        # Кнопка назад
        btn_back = tk.Button(btn_frame, text="Назад", command=self.create_main_menu, width=15)
        btn_back.pack(side=tk.RIGHT, padx=10)

    # Добавим вспомогательный метод для экспорта и показа сообщения
    def export_and_show_message(self, data, class_name, date):
        try:
            filepath = self.export_to_excel(data, class_name, date)
            messagebox.showinfo("Успех", f"Данные успешно экспортированы в файл:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать данные: {str(e)}")



if __name__ == "__main__":
    root = tk.Tk()
    app = MainApp(root)
    root.mainloop()
